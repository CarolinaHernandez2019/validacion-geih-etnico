import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Filtrar Anexo GEIH Étnico", layout="wide")

# =============================================================================
# CONFIGURACIÓN DE HOJAS A FILTRAR
# =============================================================================

HOJAS_TOTAL_NACIONAL = {
    'Total Nacional_Grupos étnicos': {
        'nombre_corto': 'TN_Grupos',
        'fila_periodos': 13,
        'descripcion': 'Indicadores por grupo étnico'
    },
    'TN_Grupos étnicos_sexo': {
        'nombre_corto': 'TN_Sexo', 
        'fila_periodos': 13,
        'descripcion': 'Indicadores por grupo étnico y sexo'
    },
    'Ocu TN_Rama': {
        'nombre_corto': 'TN_Rama',
        'fila_periodos': 12,
        'descripcion': 'Ocupados por rama de actividad'
    },
    'Ocu TN_Posocu': {
        'nombre_corto': 'TN_Posocu',
        'fila_periodos': 12,
        'descripcion': 'Ocupados por posición ocupacional'
    }
}

# Colores - Simple: Verde = bien, Rojo = mal
VERDE = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
ROJO = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
GRIS = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

borde = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# =============================================================================
# FUNCIONES
# =============================================================================

def encontrar_ultimo_periodo(df, fila_periodos):
    """Encuentra el último período de la fila"""
    ultimo_col = None
    ultimo_nombre = None
    
    for col in range(1, df.shape[1]):
        val = df.iloc[fila_periodos, col]
        if pd.notna(val) and str(val).strip():
            ultimo_col = col
            ultimo_nombre = str(val).strip()
    
    return ultimo_col, ultimo_nombre

def encontrar_columnas_mismo_patron(df, fila_periodos):
    """
    Detecta el patrón del último período (ej: Oct-Sep, Dic-Nov)
    y encuentra todas las columnas con el mismo patrón
    """
    import re
    
    ultimo_col, ultimo_nombre = encontrar_ultimo_periodo(df, fila_periodos)
    
    if not ultimo_nombre:
        return {}, None, None
    
    # Extraer patrón: "Oct 24 - Sep 25" -> mes_inicio="Oct", mes_fin="Sep"
    match = re.match(r'([A-Za-z]+)\s*\d+.*-\s*([A-Za-z]+)\s*\d+', ultimo_nombre)
    
    if not match:
        return {}, None, None
    
    mes_inicio = match.group(1)
    mes_fin = match.group(2)
    
    # Buscar todas las columnas con ese patrón
    columnas = {}
    for col in range(1, df.shape[1]):
        val = df.iloc[fila_periodos, col]
        if pd.notna(val):
            texto = str(val).strip()
            if mes_inicio in texto and mes_fin in texto:
                columnas[col] = texto
    
    return columnas, mes_inicio, mes_fin

def filtrar_hoja(df, fila_periodos, num_periodos=4):
    """
    Filtra una hoja dejando solo:
    - Columna A (conceptos)
    - Últimas N columnas del MISMO patrón de período
    """
    columnas, mes_inicio, mes_fin = encontrar_columnas_mismo_patron(df, fila_periodos)
    
    if not columnas:
        return None, [], None
    
    # Tomar las últimas N columnas del mismo patrón
    cols_ordenadas = sorted(columnas.keys())[-num_periodos:]
    nombres_periodos = [columnas[c] for c in cols_ordenadas]
    
    # Columna A + columnas de períodos
    cols_a_mantener = [0] + cols_ordenadas
    
    # Crear nuevo DataFrame
    df_filtrado = df.iloc[:, cols_a_mantener].copy()
    df_filtrado.columns = range(len(cols_a_mantener))
    
    patron = f"{mes_inicio}-{mes_fin}" if mes_inicio else None
    
    return df_filtrado, nombres_periodos, patron

def crear_excel_filtrado_simple(datos_hojas, periodos_grafico=4, periodos_tabla=2):
    """
    Crea Excel con las hojas filtradas del anexo
    Todos los datos en VERDE (luego el usuario marca en rojo los errores)
    """
    wb = Workbook()
    primera_hoja = True
    
    # Configuración de hojas - todas en verde
    config = {
        'TN_Grupos': {
            'nombre': 'H1_Grafico_4años',
            'periodos': periodos_grafico,
            'titulo_color': '375623'  # Verde oscuro
        },
        'TN_Grupos_2': {
            'nombre': 'H3_Tabla_2años',
            'periodos': periodos_tabla,
            'titulo_color': '375623'
        },
        'TN_Sexo': {
            'nombre': 'H3_Sexo',
            'periodos': periodos_tabla,
            'titulo_color': '375623'
        },
        'TN_Rama': {
            'nombre': 'H4_Rama',
            'periodos': periodos_tabla,
            'titulo_color': '375623'
        },
        'TN_Posocu': {
            'nombre': 'H5_Posocu',
            'periodos': periodos_tabla,
            'titulo_color': '375623'
        }
    }
    
    for hoja_key, hoja_config in config.items():
        # Para TN_Grupos_2, usar los datos de TN_Grupos
        hoja_datos = 'TN_Grupos' if hoja_key == 'TN_Grupos_2' else hoja_key
        
        if hoja_datos not in datos_hojas:
            continue
        
        df_original, fila_periodos = datos_hojas[hoja_datos]
        resultado = filtrar_hoja(df_original, fila_periodos, hoja_config['periodos'])
        df_filtrado, periodos = resultado[0], resultado[1]
        
        if df_filtrado is None:
            continue
        
        # Crear hoja
        if primera_hoja:
            ws = wb.active
            ws.title = hoja_config['nombre'][:31]
            primera_hoja = False
        else:
            ws = wb.create_sheet(hoja_config['nombre'][:31])
        
        # Título
        num_cols = len(periodos) + 1
        ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
        ws['A1'] = f"📊 {hoja_config['nombre']} - {', '.join(periodos)}"
        ws['A1'].font = Font(bold=True, size=11, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=hoja_config['titulo_color'], 
                                     end_color=hoja_config['titulo_color'], fill_type='solid')
        
        # Encabezados de período en fila 2
        ws.cell(row=2, column=1, value='Concepto').font = Font(bold=True)
        ws.cell(row=2, column=1).fill = GRIS
        ws.cell(row=2, column=1).border = borde
        
        for i, periodo in enumerate(periodos, 2):
            cell = ws.cell(row=2, column=i, value=periodo)
            cell.font = Font(bold=True)
            cell.fill = GRIS
            cell.border = borde
            cell.alignment = Alignment(horizontal='center')
        
        # Datos - TODO EN VERDE
        for row_idx in range(len(df_filtrado)):
            for col_idx in range(df_filtrado.shape[1]):
                valor = df_filtrado.iloc[row_idx, col_idx]
                cell = ws.cell(row=row_idx + 3, column=col_idx + 1)
                
                if pd.notna(valor):
                    if isinstance(valor, (int, float)) and col_idx > 0:
                        cell.value = round(float(valor), 1)
                        cell.fill = VERDE  # Todo verde por defecto
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        cell.value = valor
                
                cell.border = borde
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 50
        for i in range(len(periodos)):
            ws.column_dimensions[get_column_letter(i + 2)].width = 16
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# =============================================================================
# INTERFAZ
# =============================================================================

st.title("📋 Filtrar Anexo GEIH - Población Étnica")
st.markdown("""
**¿Qué hace esta app?**
1. Subes el anexo Excel
2. Detecta automáticamente el último período (Dic-Nov)
3. Filtra las hojas de Total Nacional
4. Genera un Excel con solo las columnas que necesitas para validar el boletín
""")

st.markdown("---")

# Subir archivo
uploaded_file = st.file_uploader("📂 Sube el anexo (debe tener 'anexo' en el nombre)", type=['xlsx', 'xls'])

if uploaded_file:
    try:
        xlsx = pd.ExcelFile(uploaded_file)
        st.success(f"✅ Archivo cargado: **{uploaded_file.name}**")
        
        # Mostrar hojas encontradas
        st.write("**Hojas en el archivo:**")
        
        hojas_encontradas = {}
        
        for hoja_nombre, config in HOJAS_TOTAL_NACIONAL.items():
            if hoja_nombre in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=hoja_nombre, header=None)
                columnas, mes_ini, mes_fin = encontrar_columnas_mismo_patron(df, config['fila_periodos'])
                
                if columnas:
                    ultimo_periodo = list(columnas.values())[-1]
                    patron = f"{mes_ini}-{mes_fin}"
                    hojas_encontradas[config['nombre_corto']] = (df, config['fila_periodos'])
                    st.write(f"  ✅ **{hoja_nombre}** → Patrón: **{patron}**, {len(columnas)} períodos, último: **{ultimo_periodo}**")
                else:
                    st.write(f"  ⚠️ {hoja_nombre} - No se encontraron períodos")
            else:
                st.write(f"  ❌ {hoja_nombre} - No encontrada")
        
        if hojas_encontradas:
            st.markdown("---")
            st.subheader("⚙️ Configuración de filtrado")
            
            col1, col2 = st.columns(2)
            
            with col1:
                periodos_h1 = st.selectbox(
                    "Períodos para Hoja 1 (Gráfico TD):",
                    options=[4, 3, 2, 1],
                    index=0,
                    help="Número de períodos a incluir (desde el último)"
                )
            
            with col2:
                periodos_h3 = st.selectbox(
                    "Períodos para Hoja 3 (Tablas):",
                    options=[2, 3, 4, 1],
                    index=0,
                    help="Número de períodos a incluir (desde el último)"
                )
            
            st.markdown("---")
            
            if st.button("🔄 GENERAR ANEXO FILTRADO", type="primary", use_container_width=True):
                with st.spinner("Procesando..."):
                    
                    excel_output = crear_excel_filtrado_simple(
                        hojas_encontradas, 
                        periodos_grafico=periodos_h1,
                        periodos_tabla=periodos_h3
                    )
                    
                    st.success("✅ ¡Excel generado!")
                    
                    # Preview
                    st.subheader("👀 Vista previa")
                    
                    for nombre_corto, (df, fila_per) in hojas_encontradas.items():
                        resultado = filtrar_hoja(df, fila_per, periodos_h3)
                        df_filtrado, periodos_lista = resultado[0], resultado[1]
                        if df_filtrado is not None:
                            with st.expander(f"📊 {nombre_corto} ({len(periodos_lista)} períodos)"):
                                # Renombrar columnas para mostrar
                                cols_display = ['Concepto'] + periodos_lista
                                df_display = df_filtrado.copy()
                                df_display.columns = cols_display[:df_display.shape[1]]
                                st.dataframe(df_display.head(30), use_container_width=True)
                    
                    # Botón de descarga
                    st.download_button(
                        label="📥 DESCARGAR ANEXO FILTRADO",
                        data=excel_output,
                        file_name="anexo_filtrado.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
        else:
            st.error("❌ No se encontraron hojas válidas para filtrar")
            
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        st.exception(e)

st.markdown("---")
st.markdown("""
### 📝 Hojas que se generan:

| Hoja | Para qué | Períodos |
|------|----------|----------|
| **H1_Grafico_4años** | Gráfico TD (4 años históricos) | 4 Dic-Nov |
| **H3_Tabla_2años** | Tabla 1 Total Nacional | 2 Dic-Nov |
| **H3_Sexo** | Tabla 1 por sexo | 2 Dic-Nov |
| **H4_Rama** | Rama de actividad | 2 Dic-Nov |
| **H5_Posocu** | Posición ocupacional | 2 Dic-Nov |

### 🎨 Colores:
- 🟢 **Verde** = Dato del anexo (correcto por defecto)
- 🔴 **Rojo** = Marcar manualmente si no coincide con boletín

### 📅 El filtro:
- Detecta automáticamente el último período Dic-Nov
- Elimina todas las demás columnas
- De ~121 columnas → solo 2-5 columnas
""")
